# -*- coding: utf-8 -*-
"""약품 대량등록 양식(public/templates/drug-upload-template.xlsx) 생성기.
컬럼을 바꾸려면 COLS / GUIDE / VOCAB 만 수정한 뒤 다시 실행하면 된다."""
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

P, G, L, N = '804A87', '019748', 'BFA6D9', '2E4A62'
F = '맑은 고딕'
OUT = os.path.join('public', 'templates', 'drug-upload-template.xlsx')
thin = Side(style='thin', color='D9D9D9')
BOX = Border(left=thin, right=thin, top=thin, bottom=thin)

# 19컬럼 — (라벨, 필수여부, 너비, 숫자서식, 안내문)
#   필수(보라 #804A87): 약품코드·약품명·현재고·구입단가
#   권장(라벤더 #BFA6D9): 보험코드·품목기준코드·ATC번호·규격·제형·포장·보험약가·안전재고·최대재고·보관위치
#   선택(회색): 구분·상태·유효기한·LOT번호·비고
COLS = [
    ('약품코드', '필수', 14, '@', '원내 약품코드'),
    ('약품명', '필수', 32, None, '제품명'),
    ('보험코드', '권장', 13, '@', '9자리 · 정확 매칭용'),
    ('품목기준코드', '권장', 14, '@', '9자리 · 식약처 코드'),
    ('ATC번호', '권장', 12, '@', 'WHO 해부학적 분류코드'),
    ('규격', '권장', 9, '#,##0', '입수 수량(숫자)'),
    ('제형', '권장', 10, None, '예: 정제·주사제'),
    ('포장', '권장', 10, None, '예: 병·바이알'),
    ('현재고', '필수', 10, '#,##0', '현재 재고 수량'),
    ('구입단가', '필수', 11, '#,##0', '1개당 매입가(원)'),
    ('보험약가', '권장', 11, '#,##0', '심평원 상한가'),
    ('구분', '선택', 10, None, '미입력 시 자동'),
    ('상태', '선택', 9, None, '미입력 시 사용'),
    ('안전재고', '권장', 10, '#,##0', '발주 기준점'),
    ('최대재고', '권장', 10, '#,##0', '과잉 기준점'),
    ('보관위치', '권장', 14, None, '예: A-3-2'),
    ('유효기한', '선택', 12, 'yyyy-mm-dd', '대표 로트'),
    ('LOT번호', '선택', 13, '@', '대표 로트번호'),
    ('비고', '선택', 20, None, ''),
]
# 예시 2행 — 규격=100/1 · 제형=정제/주사제 · 포장=병/바이알
EX = [
    ['ABC01', '타이레놀정500mg', '645901230', '197000123', 'N02BE01', 100, '정제', '병',
     1200, 88, 88, '경구제', '사용', 300, 2000, 'A-3-2', '2027-06-30', 'LOT2706', '예시 — 삭제하세요'],
    ['INJ07', '세프트리악손주1g', '642203450', '640012345', 'J01DD04', 1, '주사제', '바이알',
     48, 3910, 3910, '주사제', '사용', 20, 120, '냉장-B-1', '2026-11-30', 'CTX2611', '예시 — 삭제하세요'],
]
GUIDE = [
    ('약품코드', '필수', '문자', '병원에서 쓰는 원내 코드', '등록 실패'),
    ('약품명', '필수', '문자', '제품명. 보험코드가 있으면 이름이 조금 달라도 됩니다', '등록 실패'),
    ('보험코드', '권장', '9자리 숫자', '있으면 성분·ATC·보험약가가 정확히 자동 입력됩니다', '약품명으로 매칭'),
    ('품목기준코드', '권장', '9자리 숫자', '식약처 품목기준코드. 보험코드가 있으면 자동 입력됩니다', '공공DB에서 자동'),
    ('ATC번호', '권장', '영문+숫자', 'WHO 해부학적 분류코드(예: N02BE01). 보험코드가 있으면 자동', '공공DB에서 자동'),
    ('규격', '권장', '숫자', '한 포장당 입수 수량(예: 100정=100). 낱개면 1', '빈 값 유지'),
    ('제형', '권장', '문자', '예: 정제·캡슐·주사제·시럽·연고', '공공DB에서 자동'),
    ('포장', '권장', '문자', '예: 병·PTP·바이알·튜브', '공공DB에서 자동'),
    ('현재고', '필수', '숫자', '등록 시점의 재고 수량. 0도 가능', '0으로 처리'),
    ('구입단가', '필수', '숫자', '1개당 매입가(원). 부가세 포함 여부는 병원 기준대로', '재고금액이 0으로 표시됨'),
    ('보험약가', '권장', '숫자', '심평원 상한가. 구입단가 기본값=보험약가, 비급여·할인 시 직접 입력', '구입단가와 동일 처리'),
    ('구분', '선택', '목록', '경구제·주사제·외용제·영양제·수액제·의약외품', '공공DB에서 자동 입력'),
    ('상태', '선택', '목록', '사용·중지·휴면', '사용'),
    ('안전재고', '권장', '숫자', '이 수량 아래로 내려가면 발주 알림', '발주 알림 미작동'),
    ('최대재고', '권장', '숫자', '이 수량을 넘으면 과잉 표시', '과잉 판정 미작동'),
    ('보관위치', '권장', '문자', '조제실 위치. 예: A-3-2, 냉장-B-1', '위치 검색 불가'),
    ('유효기한', '선택', 'yyyy-mm-dd', '대표 로트 기준', '유효기한 알림 미작동'),
    ('LOT번호', '선택', '문자', '대표 로트번호', '—'),
    ('비고', '선택', '문자', '메모', '—'),
]
VOCAB = [
    ('구분', ['경구제', '주사제', '외용제', '영양제', '수액제', '의약외품'], '미입력 시 공공DB에서 자동'),
    ('상태', ['사용', '중지', '휴면'], '미입력 시 사용'),
    ('제형(예시)', ['정제', '캡슐', '주사제', '시럽', '연고', '점안액'], '자유 입력 — 예시 참고'),
    ('포장(예시)', ['병', 'PTP', '바이알', '앰플', '튜브', '포'], '자유 입력 — 예시 참고'),
    ('규격', ['숫자만'], '한 포장당 입수 수량. 예: 100, 30, 1'),
    ('ATC번호', ['영문+숫자 7자리'], '예: N02BE01 — 있으면 자동'),
    ('품목기준코드', ['숫자 9자리'], '식약처 코드 — 있으면 자동'),
    ('구입단가', ['숫자만'], '1개당 매입가 — 모든 금액계산 기준'),
    ('보험약가', ['숫자만'], '심평원 상한가. 구입단가 기본값=보험약가'),
    ('급여구분', ['급여', '비급여'], '자동 — 입력 불필요'),
    ('마약구분', ['일반', '향정', '마약', '한외마약'], '자동 — 입력 불필요'),
    ('보관방법', ['실온', '냉장', '실온/차광', '냉장/차광'], '자동 — 입력 불필요'),
    ('복합/단일', ['단일제', '복합제'], '자동 — 입력 불필요'),
]

def band(ws, ref, text, color, height=24, size=12):
    ws.merge_cells(ref)
    c = ws[ref.split(':')[0]]
    c.value = text
    c.font = Font(name=F, bold=True, size=size, color='FFFFFF')
    c.fill = PatternFill('solid', fgColor=color)
    c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    ws.row_dimensions[int(ref.split(':')[0][1:])].height = height

def head(ws, row, labels, color):
    for j, h in enumerate(labels, 1):
        a = ws.cell(row, j, h)
        a.font = Font(name=F, bold=True, size=10, color='FFFFFF')
        a.fill = PatternFill('solid', fgColor=color)
        a.alignment = Alignment(horizontal='center')
        a.border = BOX

wb = openpyxl.Workbook()

# ① 약품등록
ws = wb.active
ws.title = '① 약품등록'
LAST = get_column_letter(len(COLS))
band(ws, 'A1:%s1' % LAST, '약플로 약품 기초자료 등록 양식   ·   4행부터 입력하세요   ·   회색 예시 2줄은 지우고 사용하세요',
     P, height=26, size=11)
for j, (h, req, w, fmt, desc) in enumerate(COLS, 1):
    a = ws.cell(2, j, h)
    a.font = Font(name=F, bold=True, size=10, color='FFFFFF')
    a.fill = PatternFill('solid', fgColor=P if req == '필수' else (L if req == '권장' else 'BFBFBF'))
    a.alignment = Alignment(horizontal='center', vertical='center')
    a.border = BOX
    b = ws.cell(3, j, req + (' · ' + desc if desc else ''))
    b.font = Font(name=F, size=8, color=N)
    b.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    b.fill = PatternFill('solid', fgColor='F2F2F2')
    b.border = BOX
    ws.column_dimensions[get_column_letter(j)].width = w
ws.row_dimensions[2].height = 22
ws.row_dimensions[3].height = 28
for i, row in enumerate(EX, 4):
    for j, v in enumerate(row, 1):
        cc = ws.cell(i, j, v)
        cc.font = Font(name=F, size=10, italic=True, color='808080')
        cc.fill = PatternFill('solid', fgColor='F7F7F7')
        cc.border = BOX
        if COLS[j - 1][3]:
            cc.number_format = COLS[j - 1][3]
for i in range(6, 504):
    for j in range(1, len(COLS) + 1):
        cc = ws.cell(i, j)
        cc.font = Font(name=F, size=10)
        cc.border = BOX
        if COLS[j - 1][3]:
            cc.number_format = COLS[j - 1][3]
# 드롭다운: 구분·상태 (열 위치는 COLS 순서에 맞춰 자동 산출)
col_gubun = get_column_letter([c[0] for c in COLS].index('구분') + 1)
col_state = get_column_letter([c[0] for c in COLS].index('상태') + 1)
dv1 = DataValidation(type='list', formula1='"경구제,주사제,외용제,영양제,수액제,의약외품"', allow_blank=True)
dv2 = DataValidation(type='list', formula1='"사용,중지,휴면"', allow_blank=True)
ws.add_data_validation(dv1)
ws.add_data_validation(dv2)
dv1.add('%s4:%s503' % (col_gubun, col_gubun))
dv2.add('%s4:%s503' % (col_state, col_state))
ws.freeze_panes = 'A4'

# ② 입력 안내
w2 = wb.create_sheet('② 입력 안내')
band(w2, 'A1:E1', '컬럼별 입력 안내', P)
head(w2, 2, ['컬럼', '필수 여부', '형식', '설명', '미입력 시'], N)
for i, row in enumerate(GUIDE, 3):
    for j, v in enumerate(row, 1):
        cc = w2.cell(i, j, v)
        cc.font = Font(name=F, size=10)
        cc.border = BOX
        cc.alignment = Alignment(vertical='center', wrap_text=(j in (4, 5)))
        if j == 2:
            cc.font = Font(name=F, size=10, bold=(v == '필수'), color=(P if v == '필수' else N))
for j, w in enumerate([14, 10, 14, 52, 26], 1):
    w2.column_dimensions[get_column_letter(j)].width = w
r = len(GUIDE) + 4
band(w2, 'A%d:E%d' % (r, r), '자동으로 채워지는 항목 (입력하지 않으셔도 됩니다)', G, height=20, size=10)
w2.merge_cells('A%d:E%d' % (r + 1, r + 2))
c = w2.cell(r + 1, 1,
            '성분명(한글·영문) · 약효분류 · 제조사 · 단위 · 급여구분 · 전문/일반 · '
            '마약구분 · 복합/단일 · 첨가제 · 보관방법 · 효능\n'
            '→ 식약처·심평원 공공데이터에서 약플로가 자동으로 채웁니다. '
            '품목기준코드·ATC번호·제형·규격·포장도 비워 두면 보험코드 기준으로 자동 입력됩니다.')
c.font = Font(name=F, size=10, color=N)
c.alignment = Alignment(vertical='center', wrap_text=True, indent=1)
w2.row_dimensions[r + 1].height = 40

# ③ 허용값 목록
w3 = wb.create_sheet('③ 허용값 목록')
band(w3, 'A1:C1', '허용값 목록 — 아래 값 외에는 등록되지 않습니다', P)
head(w3, 2, ['항목', '허용값', '비고'], N)
i = 3
for name, vals, note in VOCAB:
    for j, v in enumerate([name, ' · '.join(vals), note], 1):
        cc = w3.cell(i, j, v)
        cc.font = Font(name=F, size=10, bold=(j == 1))
        cc.border = BOX
        cc.alignment = Alignment(vertical='center')
    i += 1
for j, w in enumerate([16, 56, 28], 1):
    w3.column_dimensions[get_column_letter(j)].width = w
i += 1
w3.merge_cells('A%d:C%d' % (i, i + 4))
c = w3.cell(i, 1,
            '주의 사항\n'
            '· 약품코드·보험코드·품목기준코드·ATC번호·LOT번호 열은 반드시 "텍스트" 서식으로 두세요.\n'
            '   숫자로 바꾸면 앞자리 0이 사라지거나 APR2 같은 코드가 날짜(2026-04-02)로 변환됩니다.\n'
            '   이 양식에는 이미 설정되어 있습니다.\n'
            '· 같은 약품코드를 두 번 넣으면 두 번째 행부터 오류로 처리됩니다.\n'
            '· 빈 칸으로 두면 기존 값이 유지됩니다. 값을 지우려면 올바른 값으로 덮어써 주세요.')
c.font = Font(name=F, size=10, color=N)
c.alignment = Alignment(vertical='top', wrap_text=True, indent=1)
c.fill = PatternFill('solid', fgColor='FDF7E3')
for k in range(5):
    w3.row_dimensions[i + k].height = 18

os.makedirs(os.path.dirname(OUT), exist_ok=True)
wb.save(OUT)
print('생성 완료:', OUT)